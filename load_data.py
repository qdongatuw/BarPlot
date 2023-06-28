import numpy as np
import os
from openpyxl import load_workbook


class LoadData:
    def __init__(self, filename=None):
        self.filename = filename
        self.book = None
        self.sheet_names = []
        self.data_list = []
        self.preview_data = []
        self.length = 0
        self.group_name = []
        self.get_sheet()

    def get_sheet(self):
        if not self.filename:
            return
        if os.path.splitext(self.filename)[1] != '.xlsx':
            raise FileExistsError
        self.book = load_workbook(self.filename)
        self.sheet_names = self.book.sheetnames

    def preview(self, sheet_name=None):
        self.preview_data.clear()
        if not sheet_name:
            return
        if sheet_name not in self.sheet_names:
            return
        sheet = self.book[sheet_name]
        rs = list(sheet.rows)
        for r in rs:
            l_ = [x.value for x in r]
            self.preview_data.append(np.asarray(l_))

    def get_groups(self, sheet_name=None):
        if sheet_name not in self.sheet_names:
            return
        if not sheet_name:
            sheet = self.book[0]
        else:
            sheet = self.book[sheet_name]

        if isinstance(sheet['A1'].value, (str, )):
            row = list(sheet.rows)[0]
            self.group_name = [x.value for x in row]

    def load_data(self, sheet_name=None, by_row=0):
        def is_num(x):
            return isinstance(x, (int, float))
        self.data_list.clear()
        if not sheet_name:
            return
        if sheet_name not in self.sheet_names:
            return
        sheet = self.book[sheet_name]
        rcs = list(sheet.rows) if by_row else list(sheet.columns)
        for rc in rcs:
            l_ = [x.value for x in rc if is_num(x.value)]
            if len(l_) > 0:
                self.data_list.append(np.asarray(l_))
        self.length = len(self.data_list)
